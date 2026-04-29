from odoo import http
from odoo.http import request
import base64
import csv
import io
import json
import openpyxl
import tempfile
import subprocess
import os

import platform
import shutil

if platform.system() == "Windows":
    LIBREOFFICE_PATH = r"C:\Program Files\LibreOffice\program\soffice.exe"
else:
    LIBREOFFICE_PATH = shutil.which("soffice") or "/usr/bin/soffice"


class AttachmentPreviewController(http.Controller):

    # =========================
    # CSV / XLSX PREVIEW
    # =========================
    @http.route('/csv/preview/<int:attachment_id>', auth='user', type='http')
    def preview_attachment(self, attachment_id):

        attachment = request.env['ir.attachment'].sudo().browse(attachment_id)

        if not attachment.exists() or not attachment.datas:
            return self._json_response({"sheets": []})

        file_data = base64.b64decode(attachment.datas)
        filename = (attachment.name or "").lower()

        try:
            file_input = io.BytesIO(file_data)

            # XLSX
            if filename.endswith(".xlsx"):

                wb = openpyxl.load_workbook(
                    file_input,
                    read_only=True,
                    data_only=True
                )

                sheets_data = []
                MAX_ROWS = 500

                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows = []

                    for i, row in enumerate(sheet.iter_rows(values_only=True)):
                        if i >= MAX_ROWS:
                            break

                        rows.append([
                            str(cell) if cell is not None else ""
                            for cell in row
                        ])

                    sheets_data.append({
                        "name": sheet_name,
                        "rows": rows
                    })

                return self._json_response({"sheets": sheets_data})

            # CSV
            elif filename.endswith(".csv"):

                file_content = file_data.decode("utf-8", errors="ignore")
                csv_file = io.StringIO(file_content)

                try:
                    dialect = csv.Sniffer().sniff(file_content[:1024]) if file_content else csv.excel
                    reader = csv.reader(csv_file, dialect)
                except Exception:
                    csv_file.seek(0)
                    reader = csv.reader(csv_file)

                rows = list(reader)

                return self._json_response({
                    "sheets": [{
                        "name": attachment.name,
                        "rows": rows
                    }]
                })

        except Exception as e:
            return self._json_response({"error": str(e)})

    # =========================
    # GENERIC PDF CONVERTER (PPTX/DOCX)
    # =========================
    def _convert_to_pdf_with_cache(self, attachment, ext):

        cache_key = f"{attachment.id}_{attachment.checksum}"
        cached_name = f"cache_{cache_key}.pdf"

        preview_cache = request.env['ir.attachment'].sudo().search([
            ('name', '=', cached_name),
        ], limit=1)

        if preview_cache:
            return base64.b64decode(preview_cache.datas)

        old_caches = request.env['ir.attachment'].sudo().search([
            ('name', 'like', f"cache_{attachment.id}_")
        ])
        old_caches.unlink()

        file_data = base64.b64decode(attachment.datas)

        with tempfile.TemporaryDirectory() as tmpdir:
            input_path = os.path.join(tmpdir, f"input.{ext}")

            with open(input_path, "wb") as f:
                f.write(file_data)

            subprocess.run([
                LIBREOFFICE_PATH,
                "--headless",
                "--convert-to", "pdf",
                "--outdir", tmpdir,
                input_path
            ], check=True)

            pdf_path = os.path.join(tmpdir, "input.pdf")

            with open(pdf_path, "rb") as f:
                pdf_data = f.read()

        request.env['ir.attachment'].sudo().create({
            'name': cached_name,
            'datas': base64.b64encode(pdf_data),
            'mimetype': 'application/pdf',
        })

        return pdf_data

    # =========================
    # PPT PREVIEW
    # =========================
    @http.route('/ppt/preview/<int:attachment_id>', auth='user', type='http')
    def ppt_preview(self, attachment_id):

        attachment = request.env['ir.attachment'].sudo().browse(attachment_id)
        if not attachment.exists():
            return request.not_found()

        pdf_data = self._convert_to_pdf_with_cache(attachment, "pptx")

        return request.make_response(pdf_data, headers=[
            ("Content-Type", "application/pdf")
        ])

    # =========================
    # DOCX PREVIEW
    # =========================
    @http.route('/docx/preview/<int:attachment_id>', auth='user', type='http')
    def docx_preview(self, attachment_id):

        attachment = request.env['ir.attachment'].sudo().browse(attachment_id)
        if not attachment.exists():
            return request.not_found()

        pdf_data = self._convert_to_pdf_with_cache(attachment, "docx")

        return request.make_response(pdf_data, headers=[
            ("Content-Type", "application/pdf")
        ])

    # =========================
    # JSON RESPONSE
    # =========================
    def _json_response(self, data):
        return request.make_response(
            json.dumps(data),
            headers=[("Content-Type", "application/json")]
        )
